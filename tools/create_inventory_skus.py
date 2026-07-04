from __future__ import annotations

import re
from collections import defaultdict
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_FILE = Path("/Users/liyao/Desktop/2026年5月点数版本.xlsx")
OUTPUT_DIR = Path("/Users/liyao/Documents/Doors and Windows Shopping wesite/output/inventory")
OUTPUT_FILE = OUTPUT_DIR / "2026年5月点数版本_SKU整理.xlsx"


HEADER_ROW = 2


COLOR_CODES = {
    "black": "BLK",
    "white": "WHT",
    "grey": "GRY",
    "gray": "GRY",
    "bronze": "BRZ",
    "brown": "BRN",
}

SHIPPED_TERMS = ("出货", "已出", "sold", "shipped")


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def compact(value: Any) -> str:
    raw = text(value)
    if not raw:
        return ""
    raw = raw.replace("\n", " ")
    raw = re.sub(r"\s+", " ", raw)
    return raw.strip()


def code(value: Any, default: str = "NA") -> str:
    raw = compact(value).upper()
    raw = raw.replace("×", "X").replace("*", "X")
    raw = re.sub(r"[^A-Z0-9]+", "", raw)
    return raw or default


def normalize_number(value: Any) -> str:
    raw = compact(value)
    if not raw:
        return ""
    try:
        number = float(raw)
    except ValueError:
        return code(raw)
    return str(int(number)) if number.is_integer() else str(number).replace(".", "P")


def normalize_color(value: Any) -> str:
    raw = compact(value)
    return COLOR_CODES.get(raw.lower(), code(raw, "CLR"))


def normalize_direction(value: Any) -> str:
    raw = compact(value)
    lowered = raw.lower()
    if lowered in {"right to left", "right-to-left"}:
        return "RTL"
    if lowered in {"left to right", "left-to-right"}:
        return "LTR"
    if lowered in {"center open", "center opening", "center"}:
        return "CTR"
    return code(raw, "DIR")


def series_from_panels(value: Any) -> str:
    raw = compact(value)
    if not raw:
        return "未填扇数"
    match = re.search(r"\d+", raw)
    if not match:
        return "未填扇数"
    panels = int(match.group())
    return "75系列" if panels % 2 == 1 else "68系列"


def series_code(series: str) -> str:
    if series.startswith("75"):
        return "75"
    if series.startswith("68"):
        return "68"
    return "SERNA"


def product_type_from_model_or_sheet(model: Any, sheet_title: str, title_row: str) -> str:
    model_text = compact(model).upper()
    if model_text.startswith("D"):
        return "D"
    if model_text.startswith("W"):
        return "W"
    source = f"{sheet_title} {title_row}".lower()
    if "window" in source or "窗" in source:
        return "W"
    return "D"


def header_map(ws) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for cell in ws[HEADER_ROW]:
        label = compact(cell.value)
        if label:
            mapping[label] = cell.column
    return mapping


def find_col(mapping: dict[str, int], candidates: tuple[str, ...]) -> int | None:
    lowered = {key.lower(): col for key, col in mapping.items()}
    for candidate in candidates:
        candidate_lower = candidate.lower()
        for key, col in lowered.items():
            if candidate_lower in key:
                return col
    return None


def latest_quantity_col(ws, quantity_col: int | None) -> int | None:
    if quantity_col is None:
        return None
    latest = quantity_col
    for col in range(quantity_col + 1, ws.max_column + 1):
        header = compact(ws.cell(HEADER_ROW, col).value)
        if re.search(r"20\d{2}", header):
            latest = col
    return latest


def note_cols(ws, remark_col: int | None, quantity_col: int | None) -> list[int]:
    if remark_col is None:
        return []
    cols = [remark_col]
    for col in range(remark_col + 1, ws.max_column + 1):
        header = compact(ws.cell(HEADER_ROW, col).value)
        if header:
            break
        if quantity_col and col > quantity_col and re.search(r"20\d{2}", compact(ws.cell(HEADER_ROW, col).value)):
            break
        cols.append(col)
    return cols


def copy_header_style(source_cell, target_cell) -> None:
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)


def valid_data_row(row: dict[str, Any]) -> bool:
    return bool(row["model"] and row["width"] and row["height"] and row["color"] and row["direction"])


def is_shipped_note(value: str) -> bool:
    lowered = compact(value).lower()
    return any(term in lowered for term in SHIPPED_TERMS)


def numeric_quantity(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(INPUT_FILE)
    group_rows: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    row_records: list[dict[str, Any]] = []

    for ws in wb.worksheets:
        if ws.title == "SKU汇总":
            continue
        mapping = header_map(ws)
        model_col = find_col(mapping, ("型号",))
        width_col = find_col(mapping, ("width", "宽度"))
        height_col = find_col(mapping, ("height", "高度"))
        panel_col = find_col(mapping, ("panel", "扇"))
        color_col = find_col(mapping, ("color", "颜色"))
        direction_col = find_col(mapping, ("direction", "方向"))
        quantity_col = find_col(mapping, ("inventory数量", "数量"))
        current_qty_col = latest_quantity_col(ws, quantity_col)
        remark_col = find_col(mapping, ("备注",))
        notes_columns = note_cols(ws, remark_col, quantity_col)
        required = (model_col, width_col, height_col, color_col, direction_col)
        if any(col is None for col in required):
            continue

        title_row = compact(ws.cell(1, 1).value)
        for row_idx in range(HEADER_ROW + 1, ws.max_row + 1):
            model = compact(ws.cell(row_idx, model_col).value)
            width = compact(ws.cell(row_idx, width_col).value)
            height = compact(ws.cell(row_idx, height_col).value)
            panels = compact(ws.cell(row_idx, panel_col).value) if panel_col else ""
            color = compact(ws.cell(row_idx, color_col).value)
            direction = compact(ws.cell(row_idx, direction_col).value)
            quantity_value = ws.cell(row_idx, current_qty_col).value if current_qty_col else None
            notes = " | ".join(
                note for note in (compact(ws.cell(row_idx, col).value) for col in notes_columns) if note
            )
            row = {
                "sheet": ws.title,
                "row": row_idx,
                "model": model,
                "width": width,
                "height": height,
                "panels": panels,
                "color": color,
                "direction": direction,
                "quantity": quantity_value,
                "sheet_product_note": title_row,
                "row_note": notes,
                "shipped": is_shipped_note(notes),
            }
            if not valid_data_row(row):
                continue
            product_type = product_type_from_model_or_sheet(model, ws.title, title_row)
            series = series_from_panels(panels)
            product_note = " / ".join(part for part in (title_row, notes) if part)
            key = (
                product_type,
                normalize_number(width),
                normalize_number(height),
                compact(panels),
                series,
                color.lower(),
                direction.lower(),
                product_note.lower(),
            )
            row["product_type"] = product_type
            row["series"] = series
            row["product_note"] = product_note
            row["key"] = key
            group_rows[key].append(row)
            row_records.append(row)

    sorted_keys = sorted(
        group_rows,
        key=lambda item: (
            item[0],
            item[4],
            item[1],
            item[2],
            item[3],
            item[5],
            item[6],
            item[7],
        ),
    )
    sku_by_key: dict[tuple[Any, ...], str] = {}
    for index, key in enumerate(sorted_keys, start=1):
        product_type, width, height, panels, series, color, direction, _ = key
        sku = "-".join(
            (
                product_type,
                series_code(series),
                f"{width}X{height}",
                f"{code(panels, 'PNA')}P",
                normalize_color(color),
                normalize_direction(direction),
                f"{index:03d}",
            )
        )
        sku_by_key[key] = sku

    for ws in wb.worksheets:
        if ws.title == "SKU汇总":
            continue
        start_col = ws.max_column + 1
        headers = ["SKU", "产品类型", "系列", "库存状态", "SKU分组依据"]
        source_header = ws.cell(HEADER_ROW, max(1, ws.max_column))
        for offset, header in enumerate(headers):
            cell = ws.cell(HEADER_ROW, start_col + offset, header)
            copy_header_style(source_header, cell)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="355C45")
        records_for_sheet = [row for row in row_records if row["sheet"] == ws.title]
        for record in records_for_sheet:
            row_idx = record["row"]
            key = record["key"]
            ws.cell(row_idx, start_col, sku_by_key[key])
            ws.cell(row_idx, start_col + 1, record["product_type"])
            ws.cell(row_idx, start_col + 2, record["series"])
            ws.cell(row_idx, start_col + 3, "已出货-不计库存" if record["shipped"] else "可计库存")
            ws.cell(row_idx, start_col + 4, record["product_note"])
        ws.freeze_panes = "A3"
        for col in range(start_col, start_col + len(headers)):
            ws.column_dimensions[get_column_letter(col)].width = 22 if col != start_col + 4 else 48
        ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(start_col + len(headers) - 1)}{ws.max_row}"

    if "SKU汇总" in wb.sheetnames:
        del wb["SKU汇总"]
    summary = wb.create_sheet("SKU汇总", 0)
    summary_headers = [
        "SKU",
        "产品类型",
        "系列",
        "宽度",
        "高度",
        "扇数",
        "颜色",
        "开门方向",
        "产品备注/分组",
        "可用库存数量",
        "出货数量",
        "原始数量合计",
        "明细行数",
        "来源工作表",
        "产品型号示例",
    ]
    summary.append(summary_headers)
    for key in sorted_keys:
        rows = group_rows[key]
        first = rows[0]
        available_qty = 0.0
        shipped_qty = 0.0
        original_qty = 0.0
        for row in rows:
            quantity = numeric_quantity(row["quantity"])
            original_qty += quantity
            if row["shipped"]:
                shipped_qty += quantity
            else:
                available_qty += quantity
        models = ", ".join(dict.fromkeys(row["model"] for row in rows if row["model"]))
        sheets = ", ".join(dict.fromkeys(row["sheet"] for row in rows))
        summary.append(
            [
                sku_by_key[key],
                first["product_type"],
                first["series"],
                first["width"],
                first["height"],
                first["panels"],
                first["color"],
                first["direction"],
                first["product_note"],
                int(available_qty) if available_qty.is_integer() else available_qty,
                int(shipped_qty) if shipped_qty.is_integer() else shipped_qty,
                int(original_qty) if original_qty.is_integer() else original_qty,
                len(rows),
                sheets,
                models[:240],
            ]
        )

    header_fill = PatternFill("solid", fgColor="355C45")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
    summary.freeze_panes = "A2"
    summary.auto_filter.ref = f"A1:{get_column_letter(summary.max_column)}{summary.max_row}"
    widths = [34, 12, 12, 10, 10, 10, 12, 18, 48, 14, 12, 14, 12, 28, 42]
    for idx, width in enumerate(widths, start=1):
        summary.column_dimensions[get_column_letter(idx)].width = width
    for row in summary.iter_rows(min_row=2, min_col=10, max_col=13):
        for cell in row:
            cell.number_format = "0"

    wb.save(OUTPUT_FILE)
    print(OUTPUT_FILE)
    print(f"sku_groups={len(sorted_keys)}")
    print(f"tagged_rows={len(row_records)}")


if __name__ == "__main__":
    main()
