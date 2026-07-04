from __future__ import annotations

import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


APP_DIR = Path(__file__).resolve().parents[1]
DB_PATH = APP_DIR / "quotes.db"
WORKBOOK_PATH = APP_DIR / "output" / "inventory" / "2026年5月点数版本_SKU整理.xlsx"
IMPORT_MARKER = "Imported from SKU workbook"


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def numeric(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def map_product_id(row: dict[str, Any]) -> str:
    source = text(row["来源工作表"])
    note = text(row["产品备注/分组"]).lower()
    product_type = text(row["产品类型"]).upper()
    series = text(row["系列"])
    panels = text(row["扇数"])

    if product_type == "W" or "窗" in source or "window" in note:
        return "BW-220"

    if "纱门" in source or "screen door" in note:
        return "BD-680-4P"

    if series.startswith("68"):
        return "BD-680-4P"

    if series.startswith("75"):
        if panels == "5":
            return "BD-620"
        if panels == "4":
            return "BD-630"
        return "BD-750-3P"

    return "BD-310"


def direction_for_system(value: Any) -> str:
    raw = text(value).lower()
    if "left to right" in raw:
        return "Stack right"
    if "right to left" in raw:
        return "Stack left"
    if "right" in raw:
        return "Stack right"
    if "left" in raw:
        return "Stack left"
    if "center" in raw:
        return "Split stack"
    return text(value)


def load_summary_rows() -> list[dict[str, Any]]:
    wb = load_workbook(WORKBOOK_PATH, data_only=True, read_only=True)
    ws = wb["SKU汇总"]
    headers = [text(cell.value) for cell in ws[1]]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values, strict=False))
        available = numeric(row.get("可用库存数量"))
        if available <= 0:
            continue
        rows.append(row)
    return rows


def init_inventory_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS inventory_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            product_id TEXT NOT NULL,
            sku TEXT,
            received_date TEXT NOT NULL,
            warehouse TEXT,
            location TEXT,
            width REAL DEFAULT 0,
            height REAL DEFAULT 0,
            color TEXT,
            glass TEXT,
            frame TEXT,
            direction TEXT,
            quantity INTEGER NOT NULL DEFAULT 0,
            reserved_quantity INTEGER NOT NULL DEFAULT 0,
            unit_cost REAL DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'Available',
            notes TEXT
        )
        """
    )


def import_rows() -> tuple[int, dict[str, int]]:
    rows = load_summary_rows()
    now = datetime.now().isoformat(timespec="seconds")
    received_date = datetime.now().date().isoformat()
    counts: dict[str, int] = {}

    with sqlite3.connect(DB_PATH) as conn:
        init_inventory_table(conn)
        conn.execute("DELETE FROM inventory_items WHERE notes LIKE ?", (f"{IMPORT_MARKER}%",))
        for row in rows:
            product_id = map_product_id(row)
            quantity = int(numeric(row["可用库存数量"]))
            notes = (
                f"{IMPORT_MARKER}. Source: {text(row['来源工作表'])}. "
                f"Group: {text(row['产品备注/分组'])}. "
                f"Models: {text(row['产品型号示例'])}. "
                f"Original qty: {text(row['原始数量合计'])}; shipped qty: {text(row['出货数量'])}."
            )
            conn.execute(
                """
                INSERT INTO inventory_items
                (created_at, updated_at, product_id, sku, received_date, warehouse, location,
                 width, height, color, glass, frame, direction, quantity, reserved_quantity,
                 unit_cost, status, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    now,
                    now,
                    product_id,
                    text(row["SKU"]),
                    received_date,
                    text(row["来源工作表"]) or "Imported inventory",
                    "",
                    numeric(row["宽度"]),
                    numeric(row["高度"]),
                    text(row["颜色"]),
                    "",
                    text(row["颜色"]),
                    direction_for_system(row["开门方向"]),
                    quantity,
                    0,
                    0.0,
                    "Available",
                    notes,
                ),
            )
            counts[product_id] = counts.get(product_id, 0) + quantity
        conn.commit()
    return len(rows), counts


def main() -> None:
    imported, counts = import_rows()
    print(f"imported_sku_rows={imported}")
    for product_id, quantity in sorted(counts.items()):
        print(f"{product_id}: {quantity}")


if __name__ == "__main__":
    main()
