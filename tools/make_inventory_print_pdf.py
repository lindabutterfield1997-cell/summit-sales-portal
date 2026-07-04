from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase.pdfmetrics import registerFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


APP_DIR = Path(__file__).resolve().parents[1]
WORKBOOK = APP_DIR / "output" / "inventory" / "2026年5月点数版本_SKU整理.xlsx"
OUTPUT = APP_DIR / "output" / "inventory" / "2026年5月点数版本_SKU汇总_打印版.pdf"


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def paragraph(value: Any, style: ParagraphStyle) -> Paragraph:
    escaped = (
        clean(value)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\n", "<br/>")
    )
    return Paragraph(escaped, style)


def main() -> None:
    registerFont(UnicodeCIDFont("STSong-Light"))
    wb = load_workbook(WORKBOOK, data_only=True, read_only=True)
    ws = wb["SKU汇总"]

    normal = ParagraphStyle(
        "NormalCN",
        fontName="STSong-Light",
        fontSize=6.2,
        leading=7.4,
        alignment=TA_LEFT,
    )
    header = ParagraphStyle(
        "HeaderCN",
        parent=normal,
        fontSize=6.4,
        leading=7.8,
        alignment=TA_CENTER,
        textColor=colors.white,
    )
    title_style = ParagraphStyle(
        "TitleCN",
        fontName="STSong-Light",
        fontSize=16,
        leading=20,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#203128"),
    )
    subtitle_style = ParagraphStyle(
        "SubtitleCN",
        fontName="STSong-Light",
        fontSize=8.5,
        leading=11,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#59635e"),
    )

    rows = list(ws.iter_rows(values_only=True))
    data = []
    for row_index, row in enumerate(rows):
        style = header if row_index == 0 else normal
        data.append([paragraph(value, style) for value in row])

    col_widths = [
        1.65 * inch,  # SKU
        0.42 * inch,  # 产品类型
        0.55 * inch,  # 系列
        0.42 * inch,  # 宽度
        0.42 * inch,  # 高度
        0.42 * inch,  # 扇数
        0.58 * inch,  # 颜色
        0.78 * inch,  # 开门方向
        2.05 * inch,  # 产品备注/分组
        0.58 * inch,  # 可用库存数量
        0.50 * inch,  # 出货数量
        0.58 * inch,  # 原始数量合计
        0.50 * inch,  # 明细行数
        0.78 * inch,  # 来源工作表
        2.35 * inch,  # 产品型号示例
    ]

    doc = SimpleDocTemplate(
        str(OUTPUT),
        pagesize=landscape(A3),
        leftMargin=0.28 * inch,
        rightMargin=0.28 * inch,
        topMargin=0.35 * inch,
        bottomMargin=0.35 * inch,
        title="SKU库存汇总打印版",
    )

    table = Table(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#355C45")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (9, 1), (12, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d9dfd8")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f9f6")]),
                ("TOPPADDING", (0, 0), (-1, -1), 2.5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )

    story = [
        Paragraph("SKU库存汇总打印版", title_style),
        Paragraph("可用库存数量已排除备注中标记为出货/已出货的记录。", subtitle_style),
        Spacer(1, 0.12 * inch),
        table,
    ]
    doc.build(story)
    print(OUTPUT)


if __name__ == "__main__":
    main()
