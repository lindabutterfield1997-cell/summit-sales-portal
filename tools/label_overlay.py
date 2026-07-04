from __future__ import annotations

import io
import json
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl
import pdfplumber
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.utils import ImageReader


EXCEL_PATH = Path(
    "/Users/liyao/Library/Containers/com.tencent.xinWeChat/Data/Library/Application Support/"
    "com.tencent.xinWeChat/2.0b4.0.9/0084ff9dfa091b9233a1432e2f680249/Message/"
    "MessageTemp/060b421eeadafe5c27688c5b44aa98fa/File/111/六月十九号杨的店铺十二单.xlsx"
)
PDF_PATH = Path(
    "/Users/liyao/Library/Containers/com.tencent.xinWeChat/Data/Library/Application Support/"
    "com.tencent.xinWeChat/2.0b4.0.9/0084ff9dfa091b9233a1432e2f680249/Message/"
    "MessageTemp/060b421eeadafe5c27688c5b44aa98fa/File/111/面单/2f5be190-3f26-4123-a10a-88181a5e0699.pdf"
)
OUT_DIR = Path("output/pdf")
OUT_PDF = OUT_DIR / "六月十九号杨的店铺十二单_已加尺码颜色面单.pdf"
OUT_JSON = OUT_DIR / "label_overlay_matches.json"


def normalize_tracking(value: object) -> str:
    return re.sub(r"\D", "", str(value or ""))


def read_cell_images() -> dict[str, bytes]:
    ns = {
        "etc": "http://www.wps.cn/officeDocument/2017/etCustomData",
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    }
    rel_key = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"

    with zipfile.ZipFile(EXCEL_PATH) as archive:
        root = ET.fromstring(archive.read("xl/cellimages.xml"))
        rels = ET.fromstring(archive.read("xl/_rels/cellimages.xml.rels"))
        rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}

        image_by_id: dict[str, bytes] = {}
        for cell_image in root.findall("etc:cellImage", ns):
            pic = cell_image.find("xdr:pic", ns)
            if pic is None:
                continue
            name_node = pic.find("xdr:nvPicPr/xdr:cNvPr", ns)
            blip_node = pic.find("xdr:blipFill/a:blip", ns)
            if name_node is None or blip_node is None:
                continue
            image_id = name_node.attrib.get("name", "")
            rel_id = blip_node.attrib.get(rel_key, "")
            target = rel_map.get(rel_id, "")
            if image_id and target and target != "NULL":
                image_by_id[image_id] = archive.read(f"xl/{target}")
        return image_by_id


def formula_image_id(value: object) -> str | None:
    match = re.search(r"ID_[A-Z0-9]+", str(value or ""))
    return match.group(0) if match else None


def read_orders() -> dict[str, dict[str, object]]:
    image_by_id = read_cell_images()
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=False)
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    attr_col = headers.index("商品属性") + 1
    tracking_col = headers.index("运单号") + 1
    front_render_col = headers.index("正面效果图") + 1
    front_pattern_col = headers.index("前图案") + 1

    orders: dict[str, dict[str, object]] = {}
    for row in range(2, ws.max_row + 1):
        tracking = normalize_tracking(ws.cell(row=row, column=tracking_col).value)
        attr = str(ws.cell(row=row, column=attr_col).value or "").strip()
        if tracking and attr:
            image_id = (
                formula_image_id(ws.cell(row=row, column=front_render_col).value)
                or formula_image_id(ws.cell(row=row, column=front_pattern_col).value)
            )
            orders[tracking] = {
                "text": re.sub(r"\s+", " ", attr.replace(",", ", ")),
                "image_id": image_id,
                "image": image_by_id.get(image_id or ""),
            }
    return orders


def extract_page_tracking_numbers() -> list[str | None]:
    page_numbers: list[str | None] = []
    tracking_re = re.compile(r"9(?:\s*\d){21}")
    with pdfplumber.open(PDF_PATH) as pdf:
        for page in pdf.pages:
            text = (page.extract_text() or "").replace("O", "0")
            candidates = [normalize_tracking(item) for item in tracking_re.findall(text)]
            page_numbers.append(candidates[-1] if candidates else None)
    if any(page_numbers):
        return page_numbers

    # This PDF is image-only, so text extraction returns no tracking numbers.
    # The mapping below was verified from rendered page crops of the tracking
    # number line on all 12 labels.
    page_numbers = [
        "9200190389427723756879",
        "9200190389427723756824",
        "9200190389427723756473",
        "9200190389427723756466",
        "9200190389427723755322",
        "9200190389427723755223",
        "9200190389427723755148",
        "9200190389427723755063",
        "9200190389427723755025",
        "9200190389427723755018",
        "9200190389427723753724",
        "9200190389427723753694",
    ]
    return page_numbers


def fit_font_size(text: str, max_width: float, base_size: int = 19) -> int:
    size = base_size
    while size > 8 and pdfmetrics.stringWidth(text, "Helvetica-Bold", size) > max_width:
        size -= 1
    return size


def build_overlay(text: str, image_data: bytes | None, width: float, height: float) -> PdfReader:
    packet = io.BytesIO()
    c = canvas.Canvas(packet, pagesize=(width, height))
    # The red rectangle in the reference sits roughly at x=141-288, y=192-261
    # in PDF points on a 298 x 420 USPS label.
    box_x = width * 0.465
    box_y = height * 0.495
    box_w = width * 0.445
    box_h = height * 0.145

    if image_data:
        image_size = min(box_h - 8, box_w * 0.45)
        c.drawImage(
            ImageReader(io.BytesIO(image_data)),
            box_x + 4,
            box_y + (box_h - image_size) / 2,
            width=image_size,
            height=image_size,
            preserveAspectRatio=True,
            anchor="c",
            mask="auto",
        )

    text_x = box_x + (box_w * 0.48 if image_data else 0)
    text_w = box_w - (box_w * 0.48 if image_data else 0) - 6
    font_size = fit_font_size(text, text_w, 17)
    c.setFont("Helvetica-Bold", font_size)
    c.setFillColorRGB(0, 0, 0)
    text_width = pdfmetrics.stringWidth(text, "Helvetica-Bold", font_size)
    c.drawString(text_x + max((text_w - text_width) / 2, 0), box_y + box_h / 2 - font_size / 3, text)
    c.save()
    packet.seek(0)
    return PdfReader(packet)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    orders = read_orders()
    page_numbers = extract_page_tracking_numbers()
    source = PdfReader(str(PDF_PATH))
    writer = PdfWriter()
    matches = []

    for index, page in enumerate(source.pages):
        width = float(page.mediabox.width)
        height = float(page.mediabox.height)
        tracking = page_numbers[index] if index < len(page_numbers) else None
        order = orders.get(tracking or "", {})
        text = str(order.get("text", ""))
        image_data = order.get("image")
        if text:
            overlay = build_overlay(text, image_data if isinstance(image_data, bytes) else None, width, height)
            page.merge_page(overlay.pages[0])
        writer.add_page(page)
        matches.append(
            {
                "page": index + 1,
                "tracking": tracking,
                "text": text,
                "image_id": order.get("image_id"),
                "has_image": bool(image_data),
                "matched": bool(text),
            }
        )

    with OUT_PDF.open("wb") as fh:
        writer.write(fh)
    OUT_JSON.write_text(json.dumps(matches, ensure_ascii=False, indent=2), encoding="utf-8")

    print(OUT_PDF.resolve())
    print(json.dumps(matches, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
